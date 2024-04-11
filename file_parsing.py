from __future__ import annotations
from openpyxl import load_workbook
from datetime import date
from datacls import PupilInformation, PupilFullInformation

from typing import TYPE_CHECKING, Any
if TYPE_CHECKING:
    pass

class ExcelParser:
    def __init__(self, file: str) -> None:
        self.workbook = load_workbook(file, read_only=False)
        self.ws = self.workbook.active

    def get_subjects(self) -> tuple[str]:
        for row in self.ws.iter_rows(min_col=8, max_row=1, values_only=True): # type: ignore
            return row # type: ignore
        return tuple()

    def get_pupils_info(self) -> tuple[PupilInformation]:
        l: list[PupilInformation] = []
        for row in self.ws.iter_rows(min_row=2, max_col=7, values_only=True): # type: ignore
            row: list[Any] = list(row) # type: ignore
            d = date(row.pop(5), row.pop(4), row.pop(3))
            row.insert(3, d)
            l.append(PupilInformation(*row))
        return tuple(l) # type: ignore
    
    def get_pupils_full_info(self) -> tuple[PupilFullInformation, ...]:
        subjects = self.get_subjects()
        pupils = list(self.get_pupils_info())

        r: list[list[int]] = []
        for row in self.ws.iter_rows(min_row=2, min_col=8, values_only=True): # type: ignore
            r.append(row) # type: ignore

        for i, pupil in enumerate(pupils):
            d: dict[str, int] = {}
            for j, subject in enumerate(subjects):
                d.update({subject: r[i][j]})
            pupil = PupilFullInformation(pupil.second_name, pupil.name, pupil.third_name, pupil.birthday, pupil.diploma_id, d)
            pupils[i] = pupil

        return tuple(pupils) # type: ignore